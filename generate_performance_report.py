# imports
from PGSQL import NFLDB_SQL
import pandas as pd
pd.set_option('future.no_silent_downcasting', True)
import warnings
warnings.simplefilter(action='ignore', category=UserWarning)
import numpy as np
import store_stats
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font,Alignment
import sys

db = NFLDB_SQL()

metrics = {
	'passing': [
		'name',
		'team',
		'opp',
		'depth',
		'pass_att',
		'performance_pass_att',
		'pass_cmp',
		'pass_yds',
		'performance_pass_yds',
		'pass_yds_per_att',
		'performance_pass_yds_per_att',
		'pass_td',
		'performance_pass_td',
		'pass_int',
		'pass_sacked',
		'pass_sacked_per_att',
		'performance_pass_sacked_per_att',
		'pass_rating',
		'performance_pass_rating',
		'performance_pct_attempts_passing',
		'performance_passing_gini',
	],
	'rushing': [
		'name',
		'team',
		'opp',
		'position',
		'depth',
		'rush_att',
		'performance_rush_att',
		'rush_yds',
		'performance_rush_yds',
		'rush_yds_per_att',
		'performance_rush_yds_per_att',
		'rush_play_utilization',
		'performance_rush_play_utilization',
		'performance_rushp_total_play_util',
		'rush_td',
		'performance_rush_td'
	],
	'receiving': [
		'name',
		'team',
		'opp',
		'position',
		'depth',
		'targets',
		'performance_targets',
		'rec',
		'performance_rec',
		'rec_per_target',
		'performance_rec_per_target',
		'rec_yds',
		'performance_rec_yds',
		'rec_yds_per_att',
		'performance_rec_yds_per_att',
		'rec_yds_per_rec',
		'performance_rec_yds_per_rec',
		'rec_td',
		'performance_rec_td',
		'rec_long',
		'performance_rec_long',
		'performance_rec_play_utilization',
		'performance_recp_total_play_util'
	]
}

def get_player_stats(position,team,depth=1,side="off",year=2024):

	cols_to_remove = [
		'id',
		'gameid',
		'team',
		'opp',
		'position',
		'depth',
		'total_play_utilization',
		'performance_total_play_utilization',
		'rush_att',
		'rush_yds',
		'rush_td',
		'targets',
		'rec',
		'rec_yds',
		'rec_td',
		'rec_long'
	]

	pp_cols = db.table_cols['passing_performance'].copy()
	rushp_cols = db.table_cols['rushing_performance'].copy()
	recp_cols = db.table_cols['receiving_performance'].copy()

	for c in cols_to_remove:
		if c in pp_cols:
			pp_cols.remove(c)
		if c in rushp_cols:
			rushp_cols.remove(c)
		if c in recp_cols:
			recp_cols.remove(c)

	pp_col_string = ','.join([f"pp.{x}" for x in pp_cols])
	rushp_col_string = ','.join([f"rushp.{x}" for x in rushp_cols])
	recp_col_string = ','.join([f"recp.{x}" for x in recp_cols])

	sql = f'''
		select
			p.name as name,
			tm2.metric_abbrev as opp,
			tot.*,
			tot.pass_sacked / NULLIF(CAST(tot.pass_att AS DOUBLE PRECISION), 0) as pass_sacked_per_att,
			tot.pass_yds / NULLIF(CAST(tot.pass_att AS DOUBLE PRECISION), 0) as pass_yds_per_att,
			{pp_col_string},
			{rushp_col_string},
			{recp_col_string},
			rushp.total_play_utilization as rushp_total_play_util,
			rushp.performance_total_play_utilization as performance_rushp_total_play_util,
			recp.total_play_utilization as recp_total_play_util,
			recp.performance_total_play_utilization as performance_recp_total_play_util
		FROM pfr_total_offense AS tot
		JOIN pfr_gameinfo AS gi ON tot.gameid = gi.gameid
		JOIN pfr_players AS p ON p.playerid = tot.playerid
		JOIN pfr_teams AS tm ON (tm.teamid = gi.home_team OR tm.teamid = gi.away_team)
		JOIN pfr_teams AS tm2 ON (tm2.teamid = gi.home_team OR tm2.teamid = gi.away_team)
		LEFT JOIN passing_performance AS pp ON pp.gameid = tot.gameid AND pp.team = tot.team
		LEFT JOIN rushing_performance AS rushp ON rushp.id = tot.id
		LEFT JOIN receiving_performance AS recp ON recp.id = tot.id
		where gi.year = {year}
		and tot.position = '{position}'
		and tot.depth = {depth}
		and tm.metric_abbrev = '{team}'
	'''

	if side == 'off':
		sql += " and tot.team = tm.metric_abbrev and tot.team != tm2.metric_abbrev"
	else:
		sql += " and tot.team != tm.metric_abbrev and tot.team = tm2.metric_abbrev"

	sql += " order by gi.week,tot.depth asc"

	results = pd.read_sql(' '.join(sql.split()),db.conn)

	if side == 'def':
		results['opp'] = team

	results.fillna(0,inplace=True)

	return results

def process_passing_stats(team1,team2,position="QB",depth=1,year=2024,last_5=True):
	team = team1
	opp = team2
	team_stats = get_player_stats(position,team)
	opp_stats = get_player_stats(position,opp,side='def')

	if last_5 == True and len(team_stats) >= 5 and len(opp_stats) >= 5:
		team_stats = team_stats[len(team_stats)-5:]
		opp_stats = opp_stats[len(opp_stats)-5:]

	title = f"{team} vs {opp} {position}{depth} Passing yards"

	df = pd.concat([team_stats,opp_stats])[metrics['passing']]

	total_pass_att = df['pass_att'].mean()
	performance_pass_att = df['performance_pass_att'].mean()
	total_pass_yds_att = df['pass_yds_per_att'].mean()
	performance_pass_yds_att = df['performance_pass_yds_per_att'].mean()
	total_pass_yds = df['pass_yds'].mean()
	performance_pass_yds = df['performance_pass_yds'].mean()
	total_projected_passing = total_pass_att * performance_pass_att * total_pass_yds_att * performance_pass_yds_att
	total_projected_passing = (total_projected_passing + (total_pass_yds * performance_pass_yds))/2

	team_pass_yds_avg = df[df['team'] == team]['pass_yds'].mean()
	matchup_pass_yds_avg = df[(df['team'] != team) & (df['opp'] == opp)]['pass_yds'].mean()

	matchup_total_pass_att = df[df['team'] == team]['pass_att'].mean()
	matchup_performance_pass_att = df[(df['team'] != team) & (df['opp'] == opp)]['performance_pass_att'].mean()
	matchup_total_pass_yds_att = df[df['team'] == team]['pass_yds_per_att'].mean()
	matchup_performance_pass_yds_att = df[(df['team'] != team) & (df['opp'] == opp)]['performance_pass_yds_per_att'].mean()
	matchup_total_pass_yds = df[df['team'] == team]['pass_yds'].mean()
	matchup_performance_pass_yds = df[(df['team'] != team) & (df['opp'] == opp)]['performance_pass_yds'].mean()
	matchup_projected_passing = matchup_total_pass_att * matchup_performance_pass_att * matchup_total_pass_yds_att * matchup_performance_pass_yds_att
	matchup_projected_passing = (matchup_projected_passing + (matchup_total_pass_yds * matchup_performance_pass_yds))/2

	report = f"""
	{title}

	TEAM AVG:        {team_pass_yds_avg}
	MATCHUP AVG:     {matchup_pass_yds_avg}
	TOTAL AVG:       {total_pass_yds}
	PROJECTION:      {total_projected_passing}
	PROJ BY MATCHUP: {matchup_projected_passing}
	\n\n
	"""

	return title,df,report

def process_rushing_stats(team1,team2,position="RB",depth=1,year=2024,last_5=True):
	team = team1
	opp = team2
	team_stats = get_player_stats(position,team,depth=depth)
	opp_stats = get_player_stats(position,opp,side='def',depth=depth)

	if last_5 == True and len(team_stats) >= 5 and len(opp_stats) >= 5:
		team_stats = team_stats[len(team_stats)-5:]
		opp_stats = opp_stats[len(opp_stats)-5:]

	title = f"{team} vs {opp} {position}{depth} Rushing yards"

	df = pd.concat([team_stats,opp_stats])[metrics['rushing']]

	total_rush_att = df['rush_att'].mean()
	performance_rush_att = df['performance_rush_att'].mean()
	total_rush_yds_att = df['rush_yds_per_att'].mean()
	performance_rush_yds_att = df['performance_rush_yds_per_att'].mean()
	total_rush_yds = df['rush_yds'].mean()
	performance_rush_yds = df['performance_rush_yds'].mean()
	total_projected_rushing = total_rush_att * performance_rush_att * total_rush_yds_att * performance_rush_yds_att
	total_projected_rushing = (total_projected_rushing + (total_rush_yds * performance_rush_yds))/2

	team_rush_yds_avg = df[df['team'] == team]['rush_yds'].mean()
	matchup_rush_yds_avg = df[(df['team'] != team) & (df['opp'] == opp)]['rush_yds'].mean()

	matchup_total_rush_att = df[df['team'] == team]['rush_att'].mean()
	matchup_performance_rush_att = df[(df['team'] != team) & (df['opp'] == opp)]['performance_rush_att'].mean()
	matchup_total_rush_yds_att = df[df['team'] == team]['rush_yds_per_att'].mean()
	matchup_performance_rush_yds_att = df[(df['team'] != team) & (df['opp'] == opp)]['performance_rush_yds_per_att'].mean()
	matchup_total_rush_yds = df[df['team'] == team]['rush_yds'].mean()
	matchup_performance_rush_yds = df[(df['team'] != team) & (df['opp'] == opp)]['performance_rush_yds'].mean()
	matchup_projected_rushing = matchup_total_rush_att * matchup_performance_rush_att * matchup_total_rush_yds_att * matchup_performance_rush_yds_att
	matchup_projected_rushing = (matchup_projected_rushing + (matchup_total_rush_yds * matchup_performance_rush_yds))/2

	report = f"""
	{title}

	TEAM AVG:        {team_rush_yds_avg}
	MATCHUP AVG:     {matchup_rush_yds_avg}
	TOTAL AVG:       {total_rush_yds}
	PROJECTED:       {total_projected_rushing}
	PROJ BY MATCHUP: {matchup_projected_rushing}
	\n\n
	"""

	return title,df,report

def process_receiving_stats(team1,team2,position="WR",depth=1,year=2024,last_5=True):
	team = team1
	opp = team2
	team_stats = get_player_stats(position,team,depth=depth)
	opp_stats = get_player_stats(position,opp,side='def',depth=depth)

	if last_5 == True and len(team_stats) >= 5 and len(opp_stats) >= 5:
		team_stats = team_stats[len(team_stats)-5:]
		opp_stats = opp_stats[len(opp_stats)-5:]

	title = f"{team} vs {opp} {position}{depth} Receiving yards/receptions"

	df = pd.concat([team_stats,opp_stats])[metrics['receiving']]

	# yards and receptions
	total_targets = df['targets'].mean()
	performance_targets = df['performance_targets'].mean()
	total_rec = df['rec'].mean()
	performance_rec = df['performance_rec'].mean()
	total_rec_per_target = df['rec_per_target'].mean()
	performance_rec_per_target = df['performance_rec_per_target'].mean()
	total_rec_yds_rec = df['rec_yds_per_rec'].mean()
	performance_rec_yds_rec = df['performance_rec_yds_per_rec'].mean()
	total_rec_yds = df['rec_yds'].mean()
	performance_rec_yds = df['performance_rec_yds'].mean()
	total_projected_receiving = total_rec * performance_rec * total_rec_yds_rec * performance_rec_yds_rec
	total_projected_receiving = (total_projected_receiving + (total_rec_yds * performance_rec_yds))/2
	total_projected_receptions = total_targets * performance_targets * total_rec_per_target * performance_rec_per_target
	total_projected_receptions = (total_projected_receptions + (total_rec * performance_rec))/2

	team_rec_avg = df[df['team'] == team]['rec'].mean()
	matchup_rec_avg = df[(df['team'] != team) & (df['opp'] == opp)]['rec'].mean()
	team_rec_yds_avg = df[df['team'] == team]['rec_yds'].mean()
	matchup_rec_yds_avg = df[(df['team'] != team) & (df['opp'] == opp)]['rec_yds'].mean()

	matchup_total_targets = df[df['team'] == team]['targets'].mean()
	matchup_performance_targets = df[(df['team'] != team) & (df['opp'] == opp)]['performance_targets'].mean()
	matchup_total_rec = df[df['team'] == team]['rec'].mean()
	matchup_performance_rec = df[(df['team'] != team) & (df['opp'] == opp)]['performance_rec'].mean()
	matchup_total_rec_per_target = df[df['team'] == team]['rec_per_target'].mean()
	matchup_performance_rec_per_target = df[(df['team'] != team) & (df['opp'] == opp)]['performance_rec_per_target'].mean()
	matchup_total_yds_rec = df[df['team'] == team]['rec_yds_per_rec'].mean()
	matchup_performance_yds_rec = df[(df['team'] != team) & (df['opp'] == opp)]['performance_rec_yds_per_rec'].mean()
	matchup_total_rec_yds = df[df['team'] == team]['rec_yds'].mean()
	matchup_performance_rec_yds = df[(df['team'] != team) & (df['opp'] == opp)]['performance_rec_yds'].mean()
	matchup_projected_receiving = matchup_total_rec * matchup_performance_rec * matchup_total_yds_rec * matchup_performance_yds_rec
	matchup_projected_receiving = (matchup_projected_receiving + (matchup_total_rec_yds * matchup_performance_rec_yds))/2
	matchup_projected_receptions = matchup_total_targets * matchup_performance_targets * matchup_total_rec_per_target * matchup_performance_rec_per_target
	matchup_projected_receptions = (matchup_projected_receptions + (matchup_total_rec * matchup_performance_rec))/2

	report = f"""
	{title}

	########## YARDS ##########
	TEAM AVG:        {team_rec_yds_avg}
	MATCHUP AVG:     {matchup_rec_yds_avg}
	TOTAL AVG:       {total_rec_yds}
	PROJECTED:       {total_projected_receiving}
	PROJ BY MATCHUP: {matchup_projected_receiving}
	\n

	##### RECEPTIONS #####
	TEAM AVG:        {team_rec_avg}
	MATCHUP AVG:     {matchup_rec_avg}
	TOTAL AVG:       {total_rec}
	PROJECTED:       {total_projected_receptions}
	PROJ BY MATCHUP: {matchup_projected_receptions}
	\n
	"""

	return title,df,report

def process_reports(team1,team2):
	titles = []
	tables = []
	reports = []

	title,df,report = process_passing_stats(team1,team2)
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_rushing_stats(team1,team2,position='QB')
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_rushing_stats(team1,team2,position='RB')
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_rushing_stats(team1,team2,position='RB',depth=2)
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_receiving_stats(team1,team2,position='RB')
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_receiving_stats(team1,team2,position='RB',depth=2)
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_receiving_stats(team1,team2,position='WR')
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_receiving_stats(team1,team2,position='WR',depth=2)
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_receiving_stats(team1,team2,position='WR',depth=3)
	titles.append(title)
	tables.append(df)
	reports.append(report)

	title,df,report = process_receiving_stats(team1,team2,position='TE')
	titles.append(title)
	tables.append(df)
	reports.append(report)

	return titles,tables,reports

def save_games_to_excel(games, filename='game_reports.xlsx'):
	"""
	Save game statistics and reports to an Excel file.
	Each game will have its own sheet containing both the DataFrame and the report text.
	"""
	# Create Excel writer object
	with pd.ExcelWriter(filename, engine='openpyxl') as writer:
		for g in games:
			# Process home team
#             df_home, report_home = process_passing_stats(g['home_team'], g['away_team'])
			titles,tables,reports = process_reports(g['home_team'],g['away_team'])
			sheet_name_home = f"{g['home_team']} vs {g['away_team']}"
			save_game_to_sheet(titles, tables, reports, sheet_name_home, writer)

			# Process away team
			titles,tables,reports = process_reports(g['away_team'], g['home_team'])
			sheet_name_away = f"{g['away_team']} vs {g['home_team']}"
			save_game_to_sheet(titles, tables, reports, sheet_name_away, writer)

def save_game_to_sheet(titles, dataframes, reports, sheet_name, writer):
	"""
	Save multiple titles, DataFrames and reports to a single sheet in the Excel file.
	Each set will be arranged as: title, DataFrame, report, followed by a blank row.

	Args:
		titles (list): List of title strings
		dataframes (list): List of pandas DataFrames
		reports (list): List of report strings
		sheet_name (str): Name for the Excel sheet
		writer: Excel writer object
	"""
	# Truncate sheet name if too long (Excel has a 31 character limit)
	sheet_name = sheet_name[:31]

	# Create a new worksheet
	workbook = writer.book
	worksheet = workbook.create_sheet(sheet_name)
	writer.sheets[sheet_name] = worksheet

	# Create title font style
	title_font = Font(size=14, bold=True, underline='single')

	current_row = 1  # Excel rows are 1-based

	# Write each set of title, DataFrame, and report
	for title, df, report in zip(titles, dataframes, reports):
		# Write title
		cell = worksheet.cell(row=current_row, column=1)
		cell.value = title
		cell.font = title_font
		current_row += 2  # Skip a row after title

		# Write DataFrame
		# Convert DataFrame to rows
		for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
			for c_idx, value in enumerate(row, 1):
				cell = worksheet.cell(row=current_row + r_idx - 1,
								   column=c_idx,
								   value=value)

		current_row += len(df.index) + 2  # Skip a row after DataFrame

		# Write report text
		for line in report.split('\n'):
			cell = worksheet.cell(row=current_row, column=1)
			cell.value = line
			cell.alignment = Alignment(wrap_text=True)
			current_row += 1

		current_row += 2  # Skip two rows between sets

	# Adjust column widths
	for column in worksheet.columns:
		max_length = 0
		column = list(column)
		for cell in column:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = (max_length + 2)
		worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

if __name__ == '__main__':
	year = sys.argv[1]
	week = sys.argv[2]
	games = store_stats.get_games(year,week)
	save_games_to_excel(games, f'week{week}_{year}_analysis.xlsx')
